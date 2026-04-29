"""
丽群帆布纺织电商统计系统
功能：读取帆布订单原始数据Excel，按尺寸分组排序，生成带小计和汇总的明细表
"""

import re
import sys
import os
import platform
import subprocess
import threading
import json
import tempfile
from datetime import datetime
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, BooleanVar, StringVar
from PIL import Image, ImageTk


APP_VERSION = "v1.8"
APP_NAME = "丽群帆布纺织电商统计系统"
APP_DISPLAY_NAME = f"{APP_NAME} {APP_VERSION}"
CONFIG_PATH = os.path.join(os.path.expanduser("~"), ".liqun_canvas_order_tool.json")
SMALL_STOCK_SIZES = ['2米*2米', '2米*3米', '2米*4米', '2米*5米', '3米*3米']
LARGE_STOCK_SIZES = [
    '2米*6米', '2米*10米', '3米*4米', '3米*5米', '3米*6米', '3米*10米',
    '4米*4米', '4米*5米', '4米*6米', '4米*8米',
    '5米*5米', '5米*6米', '5米*7米', '5米*8米', '5米*10米',
    '6米*6米', '6米*8米', '6米*10米', '7米*8米', '10米*10米',
]
ALL_STOCK_SIZES = SMALL_STOCK_SIZES + LARGE_STOCK_SIZES


def load_config():
    """读取本机配置，只保存上次选择的库存表路径等轻量信息。"""
    if not os.path.exists(CONFIG_PATH):
        return {}
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_config(config):
    """保存本机配置。"""
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def extract_size(spec_name):
    """从规格名称中提取尺寸，如 '2米*3米'，提取不到则返回 '定制'"""
    match = re.search(r'(\d+(?:\.\d+)?)\s*米?\s*\*\s*(\d+(?:\.\d+)?)\s*米?', str(spec_name))
    if match:
        w = float(match.group(1))
        h = float(match.group(2))
        w_str = f"{int(w)}米" if w == int(w) else f"{w}米"
        h_str = f"{int(h)}米" if h == int(h) else f"{h}米"
        return f"{w_str}*{h_str}"
    return "定制"


def to_number(value):
    """把 Excel 中的数量转成数字，无法识别时返回0"""
    if value is None or value == '':
        return 0
    if isinstance(value, (int, float)):
        return value
    match = re.search(r'-?\d+(?:\.\d+)?', str(value))
    return float(match.group(0)) if match else 0


def parse_size_area(size_str):
    """从尺寸字符串计算面积（平方米），定制返回0"""
    if size_str == "定制":
        return 0.0
    match = re.match(r'(\d+(?:\.\d+)?)米\*(\d+(?:\.\d+)?)米', size_str)
    if match:
        return float(match.group(1)) * float(match.group(2))
    return 0.0


def size_sort_key(size_str):
    """尺寸排序：按宽*高数值排序，定制排最后"""
    if size_str == "定制":
        return (9999, 9999)
    match = re.match(r'(\d+(?:\.\d+)?)米\*(\d+(?:\.\d+)?)米', size_str)
    if match:
        return (float(match.group(1)), float(match.group(2)))
    return (9999, 9999)


def normalize_inventory(raw_inventory):
    """整理本机保存的库存数据，返回 {尺寸: 库存数量}。"""
    inventory = {}
    for raw_size, raw_qty in (raw_inventory or {}).items():
        size = extract_size(raw_size)
        if size == "定制":
            size = str(raw_size).strip()
        qty = to_number(raw_qty)
        if qty:
            inventory[size] = inventory.get(size, 0) + qty
    return inventory


def unique_output_path(output_dir, base_name):
    """生成不冲突的 xlsx 输出路径。"""
    output_path = os.path.join(output_dir, f"{base_name}.xlsx")
    counter = 2
    while os.path.exists(output_path):
        output_path = os.path.join(output_dir, f"{base_name}_{counter}.xlsx")
        counter += 1
    return output_path


def get_printers():
    """读取系统打印机列表，失败时只返回默认打印机。"""
    printers = ["默认打印机"]
    system = platform.system()
    try:
        if system == "Windows":
            cmd = [
                "powershell",
                "-NoProfile",
                "-Command",
                "Get-Printer | Select-Object -ExpandProperty Name"
            ]
            output = subprocess.check_output(cmd, text=True, stderr=subprocess.DEVNULL)
            printers.extend([line.strip() for line in output.splitlines() if line.strip()])
        else:
            output = subprocess.check_output(["lpstat", "-a"], text=True, stderr=subprocess.DEVNULL)
            printers.extend([line.split()[0] for line in output.splitlines() if line.strip()])
    except Exception:
        pass
    return list(dict.fromkeys(printers))


def print_production_items(production_items, printer_name=None):
    """打印今日加工清单。生成专用打印文件，只打印输出结果部分。"""
    if not production_items:
        raise ValueError("没有需要打印的加工清单")

    path = create_print_workbook(production_items)
    system = platform.system()
    selected = printer_name if printer_name and printer_name != "默认打印机" else None

    if system == "Windows":
        if selected:
            subprocess.Popen([
                "powershell",
                "-NoProfile",
                "-Command",
                (
                    f"$excel = New-Object -ComObject Excel.Application; "
                    f"$excel.Visible = $false; "
                    f"$wb = $excel.Workbooks.Open('{path}'); "
                    f"$wb.Worksheets.Item(1).PrintOut($null,$null,1,$false,'{selected}'); "
                    f"$wb.Close($false); "
                    f"$excel.Quit()"
                )
            ])
        else:
            os.startfile(path, "print")
    else:
        cmd = ["lpr"]
        if selected:
            cmd.extend(["-P", selected])
        cmd.append(path)
        subprocess.Popen(cmd)


def create_print_workbook(production_items):
    """创建只包含加工清单的打印专用 Excel，并设置打印区域。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "今日加工清单"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    title_fill = PatternFill(start_color="174766", end_color="174766", fill_type="solid")
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "丽群帆布今日加工清单"
    title.font = Font(bold=True, size=18, color="FFFFFF")
    title.fill = title_fill
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["序号", "尺寸", "需加工数量", "需加工平方数"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, size=12)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    total_qty = 0
    total_area = 0
    for idx, (size, qty, area) in enumerate(production_items, 1):
        values = [idx, size, qty, round(area, 2)]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=idx + 2, column=col, value=val)
            c.font = Font(size=13)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
        total_qty += qty
        total_area += area

    total_row = len(production_items) + 3
    total_values = ["总计", "", total_qty, round(total_area, 2)]
    for col, val in enumerate(total_values, 1):
        c = ws.cell(row=total_row, column=col, value=val)
        c.font = Font(bold=True, size=13)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.print_area = f"A1:D{total_row}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.freeze_panes = "A3"

    path = os.path.join(tempfile.gettempdir(), "丽群今日加工清单_打印.xlsx")
    wb.save(path)
    return path


def process_orders(input_path, output_dir, inventory_data=None, generate_excel=True):
    """处理订单数据"""
    inventory = normalize_inventory(inventory_data)
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # 读取表头，自动匹配列位置
    header_row = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]

    col_map = {}
    for idx, name in enumerate(header_row):
        if name in ('订单号', '订单编号'):
            col_map['order_no'] = idx
        elif name in ('规格名称', '商品规格', '规格'):
            col_map['spec_name'] = idx
        elif name in ('规格编码', '编码'):
            col_map['spec_code'] = idx
        elif name in ('数量', '购买数量', '订购数量', '商品数量'):
            col_map['qty'] = idx
        elif name in ('备注', '买家备注', '卖家备注'):
            col_map['remark'] = idx
        elif name in ('买家留言',):
            col_map['buyer_msg'] = idx
        elif name in ('快递单号', '运单号', '物流单号'):
            col_map['tracking_no'] = idx

    missing = [k for k in ('order_no', 'spec_name', 'qty') if k not in col_map]
    if missing:
        col_names = {'order_no': '订单号', 'spec_name': '规格名称', 'qty': '数量'}
        raise ValueError(
            f"Excel表头中未找到必要的列：{', '.join(col_names[k] for k in missing)}\n"
            f"当前表头：{header_row}\n"
            f"请确认Excel文件格式是否正确"
        )

    orders = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        order_no = row[col_map['order_no']]
        spec_name = row[col_map['spec_name']]
        spec_code = row[col_map.get('spec_code', -1)] if col_map.get('spec_code') is not None and col_map.get('spec_code') < len(row) else ''
        qty = row[col_map['qty']]
        remark = row[col_map.get('remark', -1)] if col_map.get('remark') is not None and col_map.get('remark') < len(row) else ''
        buyer_msg = row[col_map.get('buyer_msg', -1)] if col_map.get('buyer_msg') is not None and col_map.get('buyer_msg') < len(row) else ''
        tracking_no = row[col_map.get('tracking_no', -1)] if col_map.get('tracking_no') is not None and col_map.get('tracking_no') < len(row) else ''

        if not order_no or not spec_name:
            continue

        # 跳过非帆布商品（如补收差价等）
        spec_str = str(spec_name)
        if '差价' in spec_str or '补' in spec_str:
            continue

        size = extract_size(spec_str)

        # 合并备注和买家留言
        remark_parts = []
        if remark:
            remark_parts.append(str(remark))
        if buyer_msg:
            remark_parts.append(str(buyer_msg))
        remark_text = ' | '.join(remark_parts)
        try:
            qty_val = int(qty) if qty else 0
        except (ValueError, TypeError):
            qty_val = 0

        orders.append({
            'order_no': str(order_no),
            'spec_name': str(spec_name),
            'spec_code': str(spec_code) if spec_code else '',
            'qty': qty_val,
            'tracking_no': str(tracking_no) if tracking_no else '',
            'remark': remark_text,
            'size': size,
        })

    # 按尺寸分组
    grouped = OrderedDict()
    for order in orders:
        size = order['size']
        if size not in grouped:
            grouped[size] = []
        grouped[size].append(order)

    sorted_sizes = sorted(grouped.keys(), key=size_sort_key)

    # 按面积拆分为小件组（≤10m²）和大件组（>10m²）
    small_sizes = [s for s in sorted_sizes if parse_size_area(s) <= 10]
    large_sizes = [s for s in sorted_sizes if parse_size_area(s) > 10]

    # 创建输出工作簿
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Sheet1"

    header_font = Font(bold=True, size=11)
    subtotal_font = Font(bold=True, size=11, color="000000")
    subtotal_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    section_font = Font(bold=True, size=12, color="FFFFFF")
    section_fill_small = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_fill_large = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    block_total_font = Font(bold=True, size=11, color="FFFFFF")
    block_total_fill_small = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    block_total_fill_large = PatternFill(start_color="E04040", end_color="E04040", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 表头
    headers = [
        '序号', '订单号', '规格名称', '规格编码', '数量', '快递单号', '备注', '',
        '尺寸', '订单数量', '库存数量', '需加工数量', '剩余库存', '需加工平方数'
    ]
    for col, h in enumerate(headers, 1):
        cell = out_ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        if col <= 7 or col >= 9:
            cell.border = thin_border

    current_row = 2
    seq = 1
    total_qty = 0
    total_area = 0.0
    total_need_qty = 0
    total_need_area = 0.0
    summary_data_small = []
    summary_data_large = []
    abnormal_orders = [order for order in orders if order['size'] == "定制"]

    def write_section_title(ws, row, title, fill):
        """写分区标题行"""
        c = ws.cell(row=row, column=1, value=title)
        c.font = section_font
        c.fill = fill
        c.alignment = Alignment(horizontal='left')
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).font = section_font
            ws.cell(row=row, column=col).border = thin_border
        return row + 1

    def write_block_total(ws, row, label, qty, area, font, fill):
        """写区块合计行"""
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = font
        c1.fill = fill
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).font = font
        c5 = ws.cell(row=row, column=5, value=qty)
        c5.font = font
        c5.fill = fill
        c5.border = thin_border
        c5.alignment = Alignment(horizontal='center')
        return row + 1

    def write_size_group(ws, row, seq_val, size, group, grouped_data):
        """写一个尺寸的明细和小计，返回 (new_row, new_seq, group_qty, group_area)"""
        group_qty = sum(o['qty'] for o in group)
        area_per = parse_size_area(size)
        group_area = area_per * group_qty

        for order in group:
            ws.cell(row=row, column=1, value=seq_val).border = thin_border
            ws.cell(row=row, column=2, value=order['order_no']).border = thin_border
            ws.cell(row=row, column=3, value=order['spec_name']).border = thin_border
            ws.cell(row=row, column=4, value=order['spec_code']).border = thin_border
            c5 = ws.cell(row=row, column=5, value=order['qty'])
            c5.border = thin_border
            c5.alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=6, value=order['tracking_no']).border = thin_border
            ws.cell(row=row, column=7, value=order['remark']).border = thin_border
            seq_val += 1
            row += 1

        subtotal_label = f"【{size}】小计"
        c1 = ws.cell(row=row, column=1, value=subtotal_label)
        c1.font = subtotal_font
        c1.fill = subtotal_fill
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = subtotal_fill
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).font = subtotal_font
        c5 = ws.cell(row=row, column=5, value=group_qty)
        c5.font = subtotal_font
        c5.fill = subtotal_fill
        c5.border = thin_border
        c5.alignment = Alignment(horizontal='center')
        row += 1

        return row, seq_val, group_qty, group_area

    # ── 小件区块（≤10平方米）──
    block_small_qty = 0
    block_small_area = 0.0

    if small_sizes:
        current_row = write_section_title(out_ws, current_row, "≤ 10平方米（含）", section_fill_small)

        for size in small_sizes:
            group = grouped[size]
            current_row, seq, gq, ga = write_size_group(out_ws, current_row, seq, size, group, None)
            block_small_qty += gq
            block_small_area += ga
            summary_data_small.append((size, gq, ga))

        current_row = write_block_total(
            out_ws, current_row,
            f"10平方米以下（含）合计",
            block_small_qty, block_small_area,
            block_total_font, block_total_fill_small
        )
        current_row += 1  # 空行分隔

    # ── 大件区块（>10平方米）──
    block_large_qty = 0
    block_large_area = 0.0

    if large_sizes:
        current_row = write_section_title(out_ws, current_row, "> 10平方米", section_fill_large)

        for size in large_sizes:
            group = grouped[size]
            current_row, seq, gq, ga = write_size_group(out_ws, current_row, seq, size, group, None)
            block_large_qty += gq
            block_large_area += ga
            summary_data_large.append((size, gq, ga))

        current_row = write_block_total(
            out_ws, current_row,
            f"10平方米以上合计",
            block_large_qty, block_large_area,
            block_total_font, block_total_fill_large
        )

    total_qty = block_small_qty + block_large_qty
    total_area = block_small_area + block_large_area

    # 总计行
    current_row += 1
    out_ws.cell(row=current_row, column=1, value="总计").font = Font(bold=True, size=12)
    c5 = out_ws.cell(row=current_row, column=5, value=total_qty)
    c5.font = Font(bold=True, size=12)
    c5.alignment = Alignment(horizontal='center')

    def stock_plan(size, qty):
        if size == "定制":
            return 0, 0, 0, 0
        stock_qty = inventory.get(size, 0)
        need_qty = max(qty - stock_qty, 0)
        remain_qty = max(stock_qty - qty, 0)
        need_area = need_qty * parse_size_area(size)
        return stock_qty, need_qty, remain_qty, need_area

    def write_summary_header(row, title, fill):
        labels = [title, "订单数量", "库存数量", "需加工数量", "剩余库存", "需加工平方数"]
        for col_idx, val in enumerate(labels, 9):
            c = out_ws.cell(row=row, column=col_idx, value=val)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = fill
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')
        return row + 1

    def write_summary_rows(row, summary_data, total_label, total_fill):
        block_qty = 0
        block_stock = 0
        block_need = 0
        block_remain = 0
        block_need_area = 0.0

        for size, qty, _area in summary_data:
            stock_qty, need_qty, remain_qty, need_area = stock_plan(size, qty)
            if need_qty > 0:
                production_items.append((size, need_qty, need_area))
            values = [size, qty, stock_qty, need_qty, remain_qty, round(need_area, 2)]
            for col_idx, val in enumerate(values, 9):
                c = out_ws.cell(row=row, column=col_idx, value=val)
                c.border = thin_border
                c.alignment = Alignment(horizontal='center')
                if need_qty > 0:
                    c.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            block_qty += qty
            block_stock += stock_qty
            block_need += need_qty
            block_remain += remain_qty
            block_need_area += need_area
            row += 1

        values = [total_label, block_qty, block_stock, block_need, block_remain, round(block_need_area, 2)]
        for col_idx, val in enumerate(values, 9):
            c = out_ws.cell(row=row, column=col_idx, value=val)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = total_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')

        return row + 2, block_need, block_need_area

    # 右侧汇总表（I-N列）—— 分两段，含库存和加工计划
    production_items = []
    summary_row = 2

    # 小件汇总
    if summary_data_small:
        summary_row = write_summary_header(summary_row, "≤10m²", section_fill_small)
        summary_row, need_qty, need_area = write_summary_rows(
            summary_row, summary_data_small, "小计", block_total_fill_small
        )
        total_need_qty += need_qty
        total_need_area += need_area

    # 大件汇总
    if summary_data_large:
        summary_row = write_summary_header(summary_row, ">10m²", section_fill_large)
        summary_row, need_qty, need_area = write_summary_rows(
            summary_row, summary_data_large, "小计", block_total_fill_large
        )
        total_need_qty += need_qty
        total_need_area += need_area

    # 汇总总计行
    total_stock = sum(inventory.get(size, 0) for size in sorted_sizes)
    total_remain = sum(max(inventory.get(size, 0) - sum(o['qty'] for o in grouped[size]), 0) for size in sorted_sizes)
    for col_idx, val in enumerate([
        "总计", total_qty, total_stock, total_need_qty, total_remain, round(total_need_area, 2)
    ], 9):
        c = out_ws.cell(row=summary_row, column=col_idx, value=val)
        c.font = Font(bold=True)
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

    # 今日加工清单：只列出库存不够、需要加工的尺寸和数量
    plan_ws = out_wb.create_sheet("今日加工清单")
    plan_headers = ['序号', '尺寸', '需加工数量', '需加工平方数']
    for col, h in enumerate(plan_headers, 1):
        c = plan_ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.alignment = Alignment(horizontal='center')
        c.border = thin_border

    for idx, (size, need_qty, need_area) in enumerate(production_items, 1):
        values = [idx, size, need_qty, round(need_area, 2)]
        for col, val in enumerate(values, 1):
            c = plan_ws.cell(row=idx + 1, column=col, value=val)
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')
            c.font = Font(size=13)

    total_row = len(production_items) + 2
    for col, val in enumerate(['总计', '', total_need_qty, round(total_need_area, 2)], 1):
        c = plan_ws.cell(row=total_row, column=col, value=val)
        c.font = Font(bold=True)
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

    plan_ws.column_dimensions['A'].width = 8
    plan_ws.column_dimensions['B'].width = 16
    plan_ws.column_dimensions['C'].width = 14
    plan_ws.column_dimensions['D'].width = 16
    plan_ws.freeze_panes = "A2"
    plan_ws.page_setup.orientation = "portrait"
    plan_ws.page_setup.fitToWidth = 1
    plan_ws.page_setup.fitToHeight = 0
    plan_ws.sheet_properties.pageSetUpPr.fitToPage = True

    # 库存余量：列出库存表里的全部尺寸，方便查看公司现有库存
    if inventory:
        stock_ws = out_wb.create_sheet("库存余量")
        stock_headers = ['尺寸', '库存数量', '今日订单数量', '加工后剩余库存']
        for col, h in enumerate(stock_headers, 1):
            c = stock_ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        all_stock_sizes = sorted(inventory.keys(), key=size_sort_key)
        for row_idx, size in enumerate(all_stock_sizes, 2):
            order_qty = sum(o['qty'] for o in grouped.get(size, []))
            remain_qty = max(inventory.get(size, 0) - order_qty, 0)
            values = [size, inventory.get(size, 0), order_qty, remain_qty]
            for col, val in enumerate(values, 1):
                c = stock_ws.cell(row=row_idx, column=col, value=val)
                c.border = thin_border
                c.alignment = Alignment(horizontal='center')

        stock_ws.column_dimensions['A'].width = 16
        stock_ws.column_dimensions['B'].width = 14
        stock_ws.column_dimensions['C'].width = 14
        stock_ws.column_dimensions['D'].width = 16

    if abnormal_orders:
        abnormal_ws = out_wb.create_sheet("异常订单")
        abnormal_headers = ['序号', '订单号', '规格名称', '规格编码', '数量', '快递单号', '备注']
        warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for col, h in enumerate(abnormal_headers, 1):
            c = abnormal_ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = warning_fill
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

        for idx, order in enumerate(abnormal_orders, 1):
            values = [
                idx, order['order_no'], order['spec_name'], order['spec_code'],
                order['qty'], order['tracking_no'], order['remark']
            ]
            for col, val in enumerate(values, 1):
                c = abnormal_ws.cell(row=idx + 1, column=col, value=val)
                c.border = thin_border
                c.alignment = Alignment(horizontal='center')

        abnormal_ws.column_dimensions['A'].width = 8
        abnormal_ws.column_dimensions['B'].width = 28
        abnormal_ws.column_dimensions['C'].width = 55
        abnormal_ws.column_dimensions['D'].width = 12
        abnormal_ws.column_dimensions['E'].width = 8
        abnormal_ws.column_dimensions['F'].width = 22
        abnormal_ws.column_dimensions['G'].width = 35
        abnormal_ws.freeze_panes = "A2"

    # 列宽
    out_ws.column_dimensions['A'].width = 16
    out_ws.column_dimensions['B'].width = 28
    out_ws.column_dimensions['C'].width = 55
    out_ws.column_dimensions['D'].width = 12
    out_ws.column_dimensions['E'].width = 8
    out_ws.column_dimensions['F'].width = 22
    out_ws.column_dimensions['G'].width = 35
    out_ws.column_dimensions['H'].width = 3
    out_ws.column_dimensions['I'].width = 14
    out_ws.column_dimensions['J'].width = 10
    out_ws.column_dimensions['K'].width = 12
    out_ws.column_dimensions['L'].width = 12
    out_ws.column_dimensions['M'].width = 12
    out_ws.column_dimensions['N'].width = 14

    today = datetime.now().strftime("%Y%m%d")
    base_name = f"帆布订单明细_{today}"
    output_path = unique_output_path(output_dir, base_name) if generate_excel else None

    if generate_excel:
        out_wb.save(output_path)
    return (
        output_path, len(orders), total_qty, len(sorted_sizes), round(total_area, 2),
        total_need_qty, round(total_need_area, 2), len(abnormal_orders), production_items
    )


def open_folder(path):
    """跨平台打开文件夹"""
    system = platform.system()
    if system == "Windows":
        os.startfile(path)
    elif system == "Darwin":
        subprocess.call(["open", path])
    else:
        subprocess.call(["xdg-open", path])


def resource_path(filename):
    """获取资源文件路径（兼容 PyInstaller 打包）"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)


class OrderApp:
    # 配色方案
    BG = "#edf1f5"
    CARD_BG = "#ffffff"
    PRIMARY = "#276f9f"
    PRIMARY_DARK = "#174766"
    SUCCESS = "#249447"
    DANGER = "#c7362f"
    TEXT = "#263238"
    TEXT_LIGHT = "#6b7785"
    BORDER = "#cbd5df"

    def __init__(self):
        self.root = tk.Tk()
        self.root.title(APP_DISPLAY_NAME)
        self.root.configure(bg=self.BG)
        self.root.resizable(True, True)

        # 设置窗口图标
        try:
            ico_path = resource_path("logo.ico")
            self.root.iconbitmap(ico_path)
        except Exception:
            pass

        w, h = 1080, 720
        x = (self.root.winfo_screenwidth() - w) // 2
        y = (self.root.winfo_screenheight() - h) // 2
        self.root.geometry(f"{w}x{h}+{x}+{y}")
        self.root.minsize(1020, 650)

        self.input_path = StringVar()
        self.output_dir = StringVar()
        self.inventory_status = StringVar()
        self.printer_name = StringVar(value="默认打印机")
        self.generate_excel = BooleanVar(value=True)
        self.excel_status = StringVar()
        self.output_path = None
        self.production_items = []
        self.config = load_config()
        self.config.setdefault("inventory", {})
        self._refresh_inventory_status()
        self._refresh_excel_status()

        self._setup_styles()
        self._build_ui()
        self.root.mainloop()

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")

        style.configure("Title.TLabel", background=self.BG, foreground=self.PRIMARY,
                        font=("Microsoft YaHei", 20, "bold"))
        style.configure("HeroTitle.TLabel", background=self.PRIMARY_DARK, foreground="white",
                        font=("Microsoft YaHei", 19, "bold"))
        style.configure("HeroSub.TLabel", background=self.PRIMARY_DARK, foreground="#d8edf8",
                        font=("Microsoft YaHei", 10))
        style.configure("Card.TFrame", background=self.CARD_BG)
        style.configure("CardTitle.TLabel", background=self.CARD_BG, foreground=self.TEXT,
                        font=("Microsoft YaHei", 11, "bold"))
        style.configure("Path.TEntry", font=("Microsoft YaHei", 9))
        style.configure("Clean.TCheckbutton", background=self.CARD_BG, foreground=self.TEXT,
                        font=("Microsoft YaHei", 10))

        style.configure("Primary.TButton", font=("Microsoft YaHei", 9),
                        background=self.PRIMARY, foreground="white")
        style.map("Primary.TButton",
                  background=[("active", "#1f5c86"), ("disabled", "#a0b4c8")])

        style.configure("Success.TButton", font=("Microsoft YaHei", 11, "bold"),
                        background=self.SUCCESS, foreground="white", padding=(20, 10))
        style.map("Success.TButton",
                  background=[("active", "#218838"), ("disabled", "#a0c8a0")])

        style.configure("Info.TButton", font=("Microsoft YaHei", 9),
                        background="#17a2b8", foreground="white")
        style.map("Info.TButton",
                  background=[("active", "#138496"), ("disabled", "#a0c8d0")])

        style.configure("Result.TLabel", background=self.CARD_BG, foreground=self.TEXT_LIGHT,
                        font=("Microsoft YaHei", 10), wraplength=360, justify="left")
        style.configure("Hint.TLabel", background=self.CARD_BG, foreground=self.TEXT_LIGHT,
                        font=("Microsoft YaHei", 9))

        style.configure("green.Horizontal.TProgressbar", troughcolor="#e0e0e0",
                        background=self.SUCCESS, thickness=8)

    def _make_card(self, parent, title_text, pady=(0, 8), fill="x", expand=False):
        outer = tk.Frame(parent, bg=self.BG)
        outer.pack(fill=fill, expand=expand, pady=pady)

        title = ttk.Label(outer, text=title_text, style="CardTitle.TLabel")
        title.configure(background=self.BG)
        title.pack(anchor="w", pady=(0, 4))

        card = tk.Frame(outer, bg=self.CARD_BG, highlightbackground=self.BORDER,
                        highlightthickness=1, padx=14, pady=11)
        card.pack(fill="both" if expand else "x", expand=expand)
        return card

    def _build_ui(self):
        # 顶部标题区域：logo + 文字
        header = tk.Frame(self.root, bg=self.PRIMARY_DARK)
        header.pack(fill="x")

        header_inner = tk.Frame(header, bg=self.PRIMARY_DARK)
        header_inner.pack(fill="x", padx=28, pady=18)

        try:
            logo_img = Image.open(resource_path("logo.png"))
            logo_img = logo_img.resize((52, 52), Image.LANCZOS)
            self._logo_photo = ImageTk.PhotoImage(logo_img)
            tk.Label(header_inner, image=self._logo_photo, bg=self.PRIMARY_DARK).pack(side="left", padx=(0, 12))
        except Exception:
            pass

        title_box = tk.Frame(header_inner, bg=self.PRIMARY_DARK)
        title_box.pack(side="left", fill="x", expand=True)
        ttk.Label(title_box, text=APP_DISPLAY_NAME, style="HeroTitle.TLabel").pack(anchor="w")
        ttk.Label(title_box, text="订单整理 · 库存核算 · 加工清单 · 打印", style="HeroSub.TLabel").pack(anchor="w", pady=(3, 0))

        main = tk.Frame(self.root, bg=self.BG)
        main.pack(fill="both", expand=True, padx=24, pady=16)

        left_col = tk.Frame(main, bg=self.BG, width=430)
        left_col.pack(side="left", fill="y", padx=(0, 16))
        left_col.pack_propagate(False)

        right_col = tk.Frame(main, bg=self.BG)
        right_col.pack(side="left", fill="both", expand=True)

        # 选择文件卡片
        file_card = self._make_card(left_col, "原始数据文件", pady=(0, 10))
        file_row = tk.Frame(file_card, bg=self.CARD_BG)
        file_row.pack(fill="x")
        self.file_entry = ttk.Entry(file_row, textvariable=self.input_path,
                                    state="readonly", style="Path.TEntry")
        self.file_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ttk.Button(file_row, text="选择文件", style="Primary.TButton",
                   command=self._select_file, width=10).pack(side="right")

        # 库存维护卡片
        inventory_card = self._make_card(left_col, "库存数据", pady=(0, 10))
        inventory_row = tk.Frame(inventory_card, bg=self.CARD_BG)
        inventory_row.pack(fill="x")
        self.inventory_entry = ttk.Entry(inventory_row, textvariable=self.inventory_status,
                                         state="readonly", style="Path.TEntry")
        self.inventory_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ttk.Button(inventory_row, text="录入/预览", style="Primary.TButton",
                   command=self._open_inventory_editor, width=12).pack(side="right")

        # 保存位置卡片
        save_card = self._make_card(left_col, "保存位置", pady=(0, 10))
        save_row = tk.Frame(save_card, bg=self.CARD_BG)
        save_row.pack(fill="x")
        self.save_entry = ttk.Entry(save_row, textvariable=self.output_dir,
                                    state="readonly", style="Path.TEntry")
        self.save_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ttk.Button(save_row, text="选择文件夹", style="Primary.TButton",
                   command=self._select_output_dir, width=10).pack(side="right")

        excel_row = tk.Frame(save_card, bg=self.CARD_BG)
        excel_row.pack(fill="x", pady=(8, 0))
        ttk.Checkbutton(
            excel_row,
            text="生成Excel文件到保存位置",
            variable=self.generate_excel,
            command=self._refresh_excel_status,
            style="Clean.TCheckbutton"
        ).pack(side="left")
        ttk.Label(excel_row, textvariable=self.excel_status, style="Hint.TLabel").pack(side="left", padx=(10, 0))

        # 开始处理按钮
        action_row = tk.Frame(left_col, bg=self.BG)
        action_row.pack(fill="x", pady=(8, 5))
        self.process_btn = ttk.Button(action_row, text="开始处理", style="Success.TButton",
                                      command=self._start_process, width=20)
        self.process_btn.pack(fill="x")

        # 进度条
        self.progress = ttk.Progressbar(left_col, mode="indeterminate",
                                        style="green.Horizontal.TProgressbar")
        self.progress.pack(fill="x", pady=(0, 8))

        # 结果卡片
        result_card = self._make_card(left_col, "处理结果", pady=(0, 10))
        self.result_label = ttk.Label(result_card, text="等待处理...", style="Result.TLabel")
        self.result_label.pack(fill="x")

        # 输出结果预览
        output_card = self._make_card(right_col, "输出结果", pady=(0, 0), fill="both", expand=True)
        tree_frame = tk.Frame(output_card, bg=self.CARD_BG)
        tree_frame.pack(fill="both", expand=True)
        self.output_tree = ttk.Treeview(
            tree_frame,
            columns=("size", "qty", "area"),
            show="headings",
            height=18
        )
        tree_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.output_tree.yview)
        self.output_tree.configure(yscrollcommand=tree_scroll.set)
        self.output_tree.heading("size", text="尺寸")
        self.output_tree.heading("qty", text="需加工数量")
        self.output_tree.heading("area", text="需加工平方数")
        self.output_tree.column("size", width=190, anchor="center")
        self.output_tree.column("qty", width=140, anchor="center")
        self.output_tree.column("area", width=150, anchor="center")
        self.output_tree.pack(side="left", fill="both", expand=True)
        tree_scroll.pack(side="right", fill="y")

        print_row = tk.Frame(output_card, bg=self.CARD_BG)
        print_row.pack(fill="x", pady=(8, 0))
        ttk.Label(print_row, text="打印机", style="CardTitle.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 8))
        self.printer_combo = ttk.Combobox(
            print_row,
            textvariable=self.printer_name,
            values=get_printers(),
            state="readonly",
            width=20
        )
        self.printer_combo.grid(row=0, column=1, sticky="ew")
        ttk.Button(print_row, text="刷新打印机", style="Info.TButton",
                   command=self._refresh_printers, width=10).grid(row=0, column=2, padx=(8, 0))
        self.print_btn = ttk.Button(print_row, text="打印加工清单", style="Primary.TButton",
                                    command=self._print_output, width=14, state="disabled")
        self.print_btn.grid(row=1, column=2, sticky="e", pady=(8, 0))
        self.open_btn = ttk.Button(print_row, text="打开文件夹", style="Info.TButton",
                                   command=self._open_output_folder, width=10, state="disabled")
        self.open_btn.grid(row=1, column=1, sticky="e", pady=(8, 0))
        print_row.columnconfigure(1, weight=1)

    def _select_file(self):
        path = filedialog.askopenfilename(
            title="选择帆布订单原始数据",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")],
        )
        if path:
            self.input_path.set(path)
            if not self.output_dir.get():
                self.output_dir.set(os.path.dirname(path))

    def _refresh_excel_status(self):
        if self.generate_excel.get():
            self.excel_status.set("已打勾，会生成Excel")
        else:
            self.excel_status.set("未打勾，只预览/打印")

    def _refresh_inventory_status(self):
        inventory = normalize_inventory(self.config.get("inventory", {}))
        size_count = len(inventory)
        total_qty = sum(inventory.values())
        self.inventory_status.set(f"已保存 {size_count} 种尺寸库存，共 {total_qty:g} 张")

    def _open_inventory_editor(self):
        editor = tk.Toplevel(self.root)
        editor.title("库存录入/预览")
        editor.configure(bg=self.BG)
        editor.geometry("760x640")
        editor.transient(self.root)
        editor.grab_set()

        current_inventory = normalize_inventory(self.config.get("inventory", {}))
        entries = {}

        left = tk.Frame(editor, bg=self.BG)
        left.pack(side="left", fill="both", expand=True, padx=(16, 8), pady=14)
        right = tk.Frame(editor, bg=self.BG)
        right.pack(side="right", fill="both", expand=True, padx=(8, 16), pady=14)

        ttk.Label(left, text="库存数量", style="CardTitle.TLabel").pack(anchor="w", pady=(0, 8))
        input_canvas = tk.Canvas(left, bg=self.BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(left, orient="vertical", command=input_canvas.yview)
        input_frame = tk.Frame(input_canvas, bg=self.BG)
        input_frame.bind(
            "<Configure>",
            lambda _event: input_canvas.configure(scrollregion=input_canvas.bbox("all"))
        )
        input_canvas.create_window((0, 0), window=input_frame, anchor="nw")
        input_canvas.configure(yscrollcommand=scrollbar.set)
        input_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def add_section(title, sizes, row):
            label = tk.Label(input_frame, text=title, bg=self.PRIMARY, fg="white",
                             font=("Microsoft YaHei", 10, "bold"), padx=8, pady=4)
            label.grid(row=row, column=0, columnspan=2, sticky="ew", pady=(0, 4))
            row += 1
            for size in sizes:
                tk.Label(input_frame, text=size, bg=self.BG, fg=self.TEXT,
                         font=("Microsoft YaHei", 10)).grid(row=row, column=0, sticky="w", pady=3)
                var = StringVar(value=str(current_inventory.get(size, "")))
                entry = ttk.Entry(input_frame, textvariable=var, width=10)
                entry.grid(row=row, column=1, sticky="e", padx=(12, 4), pady=3)
                entries[size] = var
                row += 1
            return row

        next_row = add_section("≤10平方米", SMALL_STOCK_SIZES, 0)
        add_section(">10平方米", LARGE_STOCK_SIZES, next_row + 1)

        ttk.Label(right, text="库存预览", style="CardTitle.TLabel").pack(anchor="w", pady=(0, 8))
        preview = ttk.Treeview(right, columns=("size", "qty"), show="headings", height=22)
        preview.heading("size", text="尺寸")
        preview.heading("qty", text="库存数量")
        preview.column("size", width=130, anchor="center")
        preview.column("qty", width=90, anchor="center")
        preview.pack(fill="both", expand=True)

        def collect_inventory():
            inventory = {}
            for size, var in entries.items():
                qty = to_number(var.get())
                if qty:
                    inventory[size] = qty
            return inventory

        def refresh_preview():
            preview.delete(*preview.get_children())
            inventory = collect_inventory()
            for size in sorted(inventory.keys(), key=size_sort_key):
                preview.insert("", "end", values=(size, f"{inventory[size]:g}"))

        def save_inventory():
            inventory = collect_inventory()
            self.config["inventory"] = inventory
            save_config(self.config)
            self._refresh_inventory_status()
            refresh_preview()
            messagebox.showinfo("完成", "库存数量已保存")

        def clear_inventory():
            for var in entries.values():
                var.set("")
            refresh_preview()

        button_row = tk.Frame(right, bg=self.BG)
        button_row.pack(fill="x", pady=(10, 0))
        ttk.Button(button_row, text="保存", style="Success.TButton",
                   command=save_inventory, width=10).pack(side="left")
        ttk.Button(button_row, text="刷新预览", style="Info.TButton",
                   command=refresh_preview, width=10).pack(side="left", padx=(8, 0))
        ttk.Button(button_row, text="清空", style="Info.TButton",
                   command=clear_inventory, width=8).pack(side="left", padx=(8, 0))
        ttk.Button(button_row, text="关闭", style="Primary.TButton",
                   command=editor.destroy, width=8).pack(side="right")

        refresh_preview()

    def _select_output_dir(self):
        initial = self.output_dir.get() or os.path.expanduser("~")
        path = filedialog.askdirectory(title="选择保存文件夹", initialdir=initial)
        if path:
            self.output_dir.set(path)

    def _start_process(self):
        if not self.input_path.get():
            messagebox.showwarning("提示", "请先选择原始数据文件")
            return
        if self.generate_excel.get() and not self.output_dir.get():
            messagebox.showwarning("提示", "请先选择保存位置")
            return

        self.process_btn.config(state="disabled")
        self.open_btn.config(state="disabled")
        self.print_btn.config(state="disabled")
        self.production_items = []
        self._populate_output_tree([])
        self.result_label.config(text="正在处理中...", foreground="gray")
        self.progress.start(15)

        thread = threading.Thread(target=self._do_process, daemon=True)
        thread.start()

    def _do_process(self):
        try:
            result = process_orders(
                self.input_path.get(),
                self.output_dir.get() or tempfile.gettempdir(),
                self.config.get("inventory", {}),
                self.generate_excel.get()
            )
            self.output_path = result[0]
            self.root.after(0, self._on_success, result)
        except Exception as e:
            self.root.after(0, self._on_error, str(e))

    def _on_success(self, result):
        (
            output_path, order_count, total_qty, size_count, total_area,
            need_qty, need_area, abnormal_count, production_items
        ) = result
        self.production_items = production_items
        self._populate_output_tree(production_items)
        abnormal_text = f"\n异常尺寸：{abnormal_count} 条（请看异常订单表）" if abnormal_count else ""
        output_text = f"已保存到：{output_path}" if output_path else "本次未生成Excel文件"
        self.progress.stop()
        self.result_label.config(
            text=(
                f"处理完成！\n"
                f"共 {order_count} 条订单    |    共 {size_count} 种尺寸\n"
                f"总数量：{total_qty}    |    总平方数：{total_area} m\u00b2\n"
                f"需加工数量：{need_qty}    |    需加工平方数：{need_area} m\u00b2\n"
                f"{output_text}"
                f"{abnormal_text}"
            ),
        )
        self.result_label.configure(foreground=self.SUCCESS)
        self.process_btn.config(state="normal")
        self.open_btn.config(state="normal" if output_path else "disabled")
        self.print_btn.config(state="normal" if production_items else "disabled")

    def _on_error(self, msg):
        self.progress.stop()
        self.result_label.config(text=f"处理失败：\n{msg}")
        self.result_label.configure(foreground=self.DANGER)
        self.process_btn.config(state="normal")

    def _populate_output_tree(self, production_items):
        self.output_tree.delete(*self.output_tree.get_children())
        for size, qty, area in production_items:
            self.output_tree.insert("", "end", values=(size, f"{qty:g}", f"{round(area, 2):g}"))

    def _refresh_printers(self):
        printers = get_printers()
        self.printer_combo.configure(values=printers)
        if self.printer_name.get() not in printers:
            self.printer_name.set("默认打印机")

    def _print_output(self):
        try:
            print_production_items(self.production_items, self.printer_name.get())
            messagebox.showinfo("完成", "加工清单已发送到打印机")
        except Exception as e:
            messagebox.showerror("打印失败", str(e))

    def _open_output_folder(self):
        if self.output_path:
            open_folder(os.path.dirname(self.output_path))


def main():
    OrderApp()


if __name__ == "__main__":
    main()
